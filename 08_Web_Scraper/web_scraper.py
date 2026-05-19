#!/usr/bin/env python3
"""
Web Scraping Modal Tool
- Paste any website URL
- Analyzes its HTML structure (like Ctrl+U)
- Shows you all available data options
- Saves your chosen data to a named Excel file
"""

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import threading
import re
from urllib.parse import urljoin, urlparse
import json


# ─────────────────────────────────────────────
#  SCRAPER CORE
# ─────────────────────────────────────────────

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}


def fetch_page(url: str):
    """Fetch URL and return BeautifulSoup object."""
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    resp = requests.get(url, headers=HEADERS, timeout=15)
    resp.raise_for_status()
    return BeautifulSoup(resp.text, "html.parser"), url


def analyze_page(soup: BeautifulSoup, base_url: str) -> dict:
    """Analyze the page and return available data categories."""
    data = {}

    # 1. Meta information
    meta = {}
    title_tag = soup.find("title")
    meta["Page Title"] = title_tag.get_text(strip=True) if title_tag else ""
    for m in soup.find_all("meta"):
        name = m.get("name") or m.get("property") or ""
        content = m.get("content") or ""
        if name and content:
            meta[name] = content
    if meta:
        data["Meta Information"] = [{"Property": k, "Value": v} for k, v in meta.items()]

    # 2. All headings
    headings = []
    for tag in ["h1", "h2", "h3", "h4", "h5", "h6"]:
        for el in soup.find_all(tag):
            text = el.get_text(strip=True)
            if text:
                headings.append({"Tag": tag.upper(), "Text": text})
    if headings:
        data["Headings (H1–H6)"] = headings

    # 3. All paragraphs
    paragraphs = []
    for p in soup.find_all("p"):
        text = p.get_text(strip=True)
        if len(text) > 20:
            paragraphs.append({"Paragraph": text})
    if paragraphs:
        data["Paragraphs / Body Text"] = paragraphs

    # 4. All links
    links = []
    seen = set()
    for a in soup.find_all("a", href=True):
        href = urljoin(base_url, a["href"])
        text = a.get_text(strip=True) or "[no text]"
        if href not in seen:
            seen.add(href)
            links.append({"Link Text": text, "URL": href})
    if links:
        data["All Hyperlinks"] = links

    # 5. Images
    images = []
    for img in soup.find_all("img"):
        src = img.get("src", "")
        if src:
            src = urljoin(base_url, src)
        alt = img.get("alt", "")
        images.append({"Alt Text": alt, "Image URL": src})
    if images:
        data["Images"] = images

    # 6. HTML Tables
    for i, table in enumerate(soup.find_all("table"), 1):
        rows_data = []
        headers_row = []
        header_cells = table.find("tr")
        if header_cells:
            headers_row = [th.get_text(strip=True) for th in header_cells.find_all(["th", "td"])]
        for row in table.find_all("tr")[1:]:
            cells = [td.get_text(strip=True) for td in row.find_all(["td", "th"])]
            if any(cells):
                row_dict = {(headers_row[j] if j < len(headers_row) else f"Col {j+1}"): v
                            for j, v in enumerate(cells)}
                rows_data.append(row_dict)
        if rows_data:
            data[f"HTML Table #{i}"] = rows_data

    # 7. List items
    list_items = []
    for ul in soup.find_all(["ul", "ol"]):
        for li in ul.find_all("li", recursive=False):
            text = li.get_text(strip=True)
            if text:
                list_items.append({"List Item": text})
    if list_items:
        data["List Items (UL/OL)"] = list_items

    # 8. Emails & Phone Numbers
    page_text = soup.get_text()
    emails = list(set(re.findall(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+", page_text)))
    phones = list(set(re.findall(r"[\+\(]?[0-9][0-9 \-\(\)]{7,}[0-9]", page_text)))
    contacts = []
    for e in emails:
        contacts.append({"Type": "Email", "Value": e})
    for p in phones:
        contacts.append({"Type": "Phone", "Value": p.strip()})
    if contacts:
        data["Emails & Phone Numbers"] = contacts

    # 9. Script / JSON-LD structured data
    json_ld_items = []
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            obj = json.loads(script.string or "")
            flat = {k: str(v) for k, v in (obj.items() if isinstance(obj, dict) else {})}
            if flat:
                json_ld_items.append(flat)
        except Exception:
            pass
    if json_ld_items:
        data["Structured Data (JSON-LD)"] = json_ld_items

    # 10. Raw HTML source snippet
    raw_html = soup.prettify()[:50000]  # limit
    lines = raw_html.splitlines()
    data["Raw HTML Source"] = [{"Line No": i + 1, "HTML": line} for i, line in enumerate(lines)]

    return data


# ─────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────

def save_to_excel(selected_data: dict, filename: str, url: str):
    """Save selected data categories to a formatted Excel file."""
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    header_fill = PatternFill("solid", start_color="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    alt_fill    = PatternFill("solid", start_color="D6E4F0")
    normal_font = Font(name="Arial", size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Cover sheet
    cover = wb.create_sheet("Summary")
    cover.column_dimensions["A"].width = 25
    cover.column_dimensions["B"].width = 60
    cover["A1"] = "Web Scraping Report"
    cover["A1"].font = Font(bold=True, size=16, color="1F4E79", name="Arial")
    cover.merge_cells("A1:B1")
    cover["A2"] = "Source URL:"
    cover["B2"] = url
    cover["A3"] = "Sheets Exported:"
    cover["B3"] = ", ".join(selected_data.keys())
    cover["A4"] = "Total Categories:"
    cover["B4"] = len(selected_data)
    for r in range(1, 5):
        cover.row_dimensions[r].height = 22

    # Data sheets
    for category, rows in selected_data.items():
        safe_name = re.sub(r'[\\/*?:\[\]]', '', category)[:31]
        ws = wb.create_sheet(safe_name)

        if not rows:
            ws["A1"] = "No data found."
            continue

        cols = list(rows[0].keys())

        # Header row
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Data rows
        for ri, row in enumerate(rows, 2):
            fill = alt_fill if ri % 2 == 0 else PatternFill("solid", start_color="FFFFFF")
            for ci, col in enumerate(cols, 1):
                val = row.get(col, "")
                cell = ws.cell(row=ri, column=ci, value=str(val) if val else "")
                cell.font = normal_font
                cell.fill = fill
                cell.alignment = left_align
                cell.border = thin_border

        # Auto column width
        for ci, col in enumerate(cols, 1):
            max_len = len(col)
            for row in rows[:50]:
                val = str(row.get(col, ""))
                max_len = max(max_len, min(len(val), 80))
            ws.column_dimensions[get_column_letter(ci)].width = max_len + 4

        ws.freeze_panes = "A2"

    wb.save(filename)
    return filename


# ─────────────────────────────────────────────
#  GUI  (Tkinter Modal)
# ─────────────────────────────────────────────

class WebScraperApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🌐 Web Scraper Tool")
        self.root.geometry("800x700")
        self.root.resizable(True, True)
        self.root.configure(bg="#f0f4f8")

        self.scraped_data = {}
        self.check_vars = {}

        self._build_ui()

    def _build_ui(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TButton", font=("Arial", 10, "bold"), padding=8)
        style.configure("TLabel",  font=("Arial", 10), background="#f0f4f8")
        style.configure("Header.TLabel", font=("Arial", 16, "bold"),
                        foreground="#1F4E79", background="#f0f4f8")
        style.configure("Sub.TLabel", font=("Arial", 9), foreground="#555",
                        background="#f0f4f8")
        style.configure("TCheckbutton", font=("Arial", 10), background="#f0f4f8")

        # ── Title
        ttk.Label(self.root, text="🌐 Web Scraper Modal",
                  style="Header.TLabel").pack(pady=(20, 2))
        ttk.Label(self.root,
                  text="Paste any website URL → analyze → pick data → export to Excel",
                  style="Sub.TLabel").pack(pady=(0, 15))

        # ── URL Frame
        url_frame = tk.Frame(self.root, bg="#f0f4f8")
        url_frame.pack(fill="x", padx=30)
        ttk.Label(url_frame, text="Website URL:").pack(anchor="w")
        entry_row = tk.Frame(url_frame, bg="#f0f4f8")
        entry_row.pack(fill="x", pady=5)
        self.url_var = tk.StringVar()
        self.url_entry = ttk.Entry(entry_row, textvariable=self.url_var,
                                   font=("Arial", 11), width=60)
        self.url_entry.pack(side="left", fill="x", expand=True, ipady=6)
        self.url_entry.insert(0, "https://")
        self.analyze_btn = ttk.Button(entry_row, text="🔍 Analyze",
                                       command=self._start_analyze)
        self.analyze_btn.pack(side="left", padx=(8, 0))

        # ── Status bar
        self.status_var = tk.StringVar(value="Enter a URL and click Analyze.")
        self.status_lbl = ttk.Label(self.root, textvariable=self.status_var,
                                    foreground="#1a6e3c", font=("Arial", 9, "italic"),
                                    background="#f0f4f8")
        self.status_lbl.pack(anchor="w", padx=30)

        # ── Progress bar
        self.progress = ttk.Progressbar(self.root, mode="indeterminate", length=300)
        self.progress.pack(pady=5)

        # ── Options frame (scrollable)
        ttk.Label(self.root, text="Available Data Categories  (check what you want to export):",
                  font=("Arial", 10, "bold"), background="#f0f4f8").pack(anchor="w", padx=30, pady=(10, 2))

        canvas_frame = tk.Frame(self.root, bg="#ffffff", relief="sunken", bd=1)
        canvas_frame.pack(fill="both", expand=True, padx=30, pady=5)

        self.canvas = tk.Canvas(canvas_frame, bg="#ffffff", highlightthickness=0)
        scrollbar = ttk.Scrollbar(canvas_frame, orient="vertical",
                                  command=self.canvas.yview)
        self.scrollable_frame = tk.Frame(self.canvas, bg="#ffffff")

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        self.canvas.configure(yscrollcommand=scrollbar.set)
        self.canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Mouse wheel scroll
        self.canvas.bind_all("<MouseWheel>",
                             lambda e: self.canvas.yview_scroll(-1*(e.delta//120), "units"))

        self._placeholder_label = ttk.Label(
            self.scrollable_frame,
            text="(Waiting for analysis…)",
            foreground="#aaa", font=("Arial", 9, "italic"), background="#ffffff"
        )
        self._placeholder_label.pack(padx=20, pady=20)

        # ── Bottom buttons
        btn_frame = tk.Frame(self.root, bg="#f0f4f8")
        btn_frame.pack(fill="x", padx=30, pady=10)

        ttk.Button(btn_frame, text="✅ Select All",
                   command=self._select_all).pack(side="left", padx=4)
        ttk.Button(btn_frame, text="❌ Deselect All",
                   command=self._deselect_all).pack(side="left", padx=4)

        self.export_btn = ttk.Button(btn_frame, text="💾 Export to Excel",
                                      command=self._export, state="disabled")
        self.export_btn.pack(side="right", padx=4)

        ttk.Button(btn_frame, text="🔄 Reset",
                   command=self._reset).pack(side="right", padx=4)

    # ── Analyze
    def _start_analyze(self):
        url = self.url_var.get().strip()
        if not url or url == "https://":
            messagebox.showwarning("Missing URL", "Please enter a valid website URL.")
            return
        self.analyze_btn.config(state="disabled")
        self.export_btn.config(state="disabled")
        self.status_var.set("Fetching & analyzing page…")
        self.progress.start(10)
        threading.Thread(target=self._analyze_thread, args=(url,), daemon=True).start()

    def _analyze_thread(self, url):
        try:
            soup, final_url = fetch_page(url)
            data = analyze_page(soup, final_url)
            self.scraped_data = data
            self.root.after(0, lambda: self._populate_options(data))
            self.root.after(0, lambda: self.status_var.set(
                f"✅ Analysis complete — {len(data)} data categories found for: {final_url}"))
        except Exception as e:
            self.root.after(0, lambda: self.status_var.set(f"❌ Error: {e}"))
            self.root.after(0, lambda: messagebox.showerror("Scraping Error", str(e)))
        finally:
            self.root.after(0, self.progress.stop)
            self.root.after(0, lambda: self.analyze_btn.config(state="normal"))

    def _populate_options(self, data):
        # Clear old widgets
        for w in self.scrollable_frame.winfo_children():
            w.destroy()
        self.check_vars.clear()

        if not data:
            ttk.Label(self.scrollable_frame, text="No data found on this page.",
                      background="#ffffff").pack(padx=20, pady=20)
            return

        colors = ["#EBF5FB", "#FDFEFE"]
        for i, (category, rows) in enumerate(data.items()):
            row_count = len(rows)
            bg = colors[i % 2]
            frame = tk.Frame(self.scrollable_frame, bg=bg, pady=4)
            frame.pack(fill="x", padx=2, pady=1)

            var = tk.BooleanVar(value=True)
            self.check_vars[category] = var

            cb = tk.Checkbutton(
                frame, text=f"  {category}", variable=var,
                bg=bg, activebackground=bg,
                font=("Arial", 10, "bold"), fg="#1F4E79", anchor="w"
            )
            cb.pack(side="left", fill="x", expand=True)

            preview = f"{row_count} rows"
            if rows:
                cols = list(rows[0].keys())
                preview += f" · Columns: {', '.join(cols[:4])}"
                if len(cols) > 4:
                    preview += f" +{len(cols)-4} more"

            tk.Label(frame, text=preview, bg=bg,
                     font=("Arial", 8), fg="#555").pack(side="right", padx=10)

        self.export_btn.config(state="normal")

    # ── Helpers
    def _select_all(self):
        for v in self.check_vars.values():
            v.set(True)

    def _deselect_all(self):
        for v in self.check_vars.values():
            v.set(False)

    def _reset(self):
        self.scraped_data.clear()
        self.check_vars.clear()
        self.url_var.set("https://")
        self.status_var.set("Enter a URL and click Analyze.")
        for w in self.scrollable_frame.winfo_children():
            w.destroy()
        self._placeholder_label = ttk.Label(
            self.scrollable_frame, text="(Waiting for analysis…)",
            foreground="#aaa", font=("Arial", 9, "italic"), background="#ffffff"
        )
        self._placeholder_label.pack(padx=20, pady=20)
        self.export_btn.config(state="disabled")

    def _export(self):
        selected = {cat: self.scraped_data[cat]
                    for cat, var in self.check_vars.items() if var.get()}
        if not selected:
            messagebox.showwarning("Nothing Selected",
                                   "Please check at least one data category.")
            return

        filename = simpledialog.askstring(
            "Save Excel File",
            "Enter a name for your Excel file (without .xlsx):",
            initialvalue="scraped_data"
        )
        if not filename:
            return
        filename = re.sub(r'[\\/*?:<>|]', '_', filename.strip())

        try:
            self.status_var.set("Saving Excel file…")
            path = save_to_excel(selected, filename, self.url_var.get())
            self.status_var.set(f"✅ Saved: {path}")
            messagebox.showinfo("Export Successful",
                                f"Data saved to:\n{path}\n\n"
                                f"Sheets: {', '.join(selected.keys())}")
        except Exception as e:
            messagebox.showerror("Export Error", str(e))
            self.status_var.set(f"❌ Export error: {e}")


# ─────────────────────────────────────────────
#  ENTRY POINT
# ─────────────────────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    app = WebScraperApp(root)
    root.mainloop()